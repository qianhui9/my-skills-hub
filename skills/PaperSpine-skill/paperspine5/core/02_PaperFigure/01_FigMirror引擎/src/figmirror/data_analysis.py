"""Deterministic tabular-study profiling and clinical CBC figure rendering.

The module adds a data-understanding layer in front of FigMirror's existing
vector export/finalization primitives.  It keeps the source workbook read-only,
uses source-supplied abnormal flags instead of inventing clinical thresholds,
and writes aggregate analysis artifacts that can be audited independently.
"""

from __future__ import annotations

import json
import math
import re
import sys
from pathlib import Path
from typing import Any

from .data import PUBLICATION_RC, sha256_file
from .evidence_bridge import build_data_evidence_bundle
from .svg import audit_svg

DATA_STUDY_DOMAINS = {"clinical-cbc"}
DEFAULT_GROUP_ORDER = ("体检", "门诊", "急诊", "住院")
GROUP_LABELS = {
    "体检": "Health exam",
    "门诊": "Outpatient",
    "急诊": "Emergency",
    "住院": "Inpatient",
}
GROUP_COLORS = {
    "体检": "#4C9F70",
    "门诊": "#4477AA",
    "急诊": "#D55E00",
    "住院": "#7A5195",
}
METRIC_ORDER = (
    "WBC",
    "NEUT%",
    "NEUT#",
    "LYM%",
    "LYM#",
    "MONO%",
    "MONO#",
    "EOSIN%",
    "EOSIN#",
    "BASO%",
    "BASO#",
    "RBC",
    "Hb",
    "HCT",
    "MCV",
    "MCH",
    "MCHC",
    "RDW-CV",
    "PLT",
    "MPV",
    "PCT",
    "PDW",
)


class DataStudyError(ValueError):
    """Raised when a tabular source cannot support the requested study."""


def _libraries() -> tuple[Any, Any]:
    try:
        import numpy as np

        # Pandas treats these as optional accelerators.  Marking them absent
        # avoids loading binary wheels that may target a different NumPy ABI;
        # none of the deterministic analysis below depends on either package.
        sys.modules.setdefault("numexpr", None)
        sys.modules.setdefault("bottleneck", None)
        import pandas as pd
    except ImportError as exc:  # pragma: no cover - host dependency failure
        raise RuntimeError("clinical data studies require numpy and pandas") from exc
    return np, pd


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    text = re.sub(r"\s+", " ", str(value)).strip()
    return "" if text.lower() in {"nan", "none", "<na>"} else text


def _is_unnamed(column: object) -> bool:
    text = _clean_text(column)
    return not text or text.lower().startswith("unnamed:")


def _load_table(source: Path, sheet_name: str | None) -> tuple[Any, str | None]:
    _, pd = _libraries()
    suffix = source.suffix.lower()
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - host dependency failure
            raise RuntimeError(".xlsx data studies require openpyxl") from exc
        workbook = load_workbook(source, read_only=True, data_only=True)
        resolved_sheet = sheet_name or str(workbook.sheetnames[0])
        if resolved_sheet not in workbook.sheetnames:
            workbook.close()
            raise DataStudyError(f"sheet not found: {resolved_sheet}")
        worksheet = workbook[resolved_sheet]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration as exc:
            workbook.close()
            raise DataStudyError("source workbook sheet is empty") from exc
        frame = pd.DataFrame(rows, columns=headers)
        workbook.close()
        return frame, resolved_sheet
    if suffix == ".xls":
        frame = pd.read_excel(source, sheet_name=sheet_name or 0, dtype=str)
        return frame, sheet_name
    if suffix == ".csv":
        return pd.read_csv(source, dtype=str), None
    if suffix == ".tsv":
        return pd.read_csv(source, sep="\t", dtype=str), None
    raise DataStudyError("supported tabular sources are .xlsx, .xls, .csv, and .tsv")


def _normalize_frame(frame: Any) -> Any:
    _, pd = _libraries()
    normalized = frame.copy()
    normalized.columns = [_clean_text(column) or f"Unnamed: {index}" for index, column in enumerate(frame.columns)]
    for column in normalized.columns:
        normalized[column] = normalized[column].map(_clean_text).replace("", pd.NA)
    normalized = normalized.dropna(how="all").reset_index(drop=True)
    return normalized


def _flag_direction(value: object) -> str:
    text = _clean_text(value)
    if "↑" in text:
        return "high"
    if "↓" in text:
        return "low"
    return "normal" if not text else "unknown"


def _infer_metric_pairs(frame: Any, metadata_columns: int) -> list[dict[str, Any]]:
    _, pd = _libraries()
    columns = list(frame.columns)
    pairs: list[dict[str, Any]] = []
    index = metadata_columns
    while index < len(columns):
        value_column = str(columns[index])
        if _is_unnamed(value_column):
            index += 1
            continue
        values = pd.to_numeric(frame[value_column], errors="coerce")
        non_missing = frame[value_column].notna().sum()
        numeric_ratio = float(values.notna().sum() / non_missing) if non_missing else 0.0
        flag_column: str | None = None
        if index + 1 < len(columns) and _is_unnamed(columns[index + 1]):
            candidate = str(columns[index + 1])
            raw_flags = frame[candidate].dropna().map(_clean_text)
            flag_like = raw_flags.map(lambda item: "↑" in item or "↓" in item)
            if raw_flags.empty or bool(flag_like.all()):
                flag_column = candidate
        if numeric_ratio >= 0.80:
            pairs.append({"metric": value_column, "value_column": value_column, "flag_column": flag_column})
            index += 2 if flag_column else 1
        else:
            index += 1
    if len(pairs) < 2:
        raise DataStudyError("could not infer at least two numeric metric columns")
    return pairs


def _parse_age_years(value: object) -> float:
    text = _clean_text(value)
    if not text:
        return float("nan")
    match = re.search(r"(-?\d+(?:\.\d+)?)\s*(岁|月|天)", text)
    if match:
        amount = float(match.group(1))
        return amount if match.group(2) == "岁" else amount / 12 if match.group(2) == "月" else amount / 365.25
    try:
        return float(text)
    except ValueError:
        return float("nan")


def _ordered_groups(series: Any, requested: list[str] | None) -> list[str]:
    present = [_clean_text(item) for item in series.dropna().unique() if _clean_text(item)]
    if requested:
        order = [item for item in requested if item in present]
        return [*order, *(item for item in present if item not in order)]
    preferred = [item for item in DEFAULT_GROUP_ORDER if item in present]
    return [*preferred, *(item for item in present if item not in preferred)]


def _benjamini_hochberg(p_values: list[float | None]) -> list[float | None]:
    indexed = [(index, float(value)) for index, value in enumerate(p_values) if value is not None and math.isfinite(value)]
    if not indexed:
        return [None for _ in p_values]
    indexed.sort(key=lambda item: item[1])
    count = len(indexed)
    adjusted: dict[int, float] = {}
    running = 1.0
    for rank, (index, value) in reversed(list(enumerate(indexed, start=1))):
        running = min(running, value * count / rank)
        adjusted[index] = min(1.0, running)
    return [adjusted.get(index) for index in range(len(p_values))]


def _regularized_gamma_q(shape: float, value: float) -> float:
    """Regularized upper incomplete gamma, adapted from standard series/CF forms."""

    if shape <= 0 or value < 0:
        raise ValueError("gamma arguments must be non-negative with positive shape")
    if value == 0:
        return 1.0
    epsilon = 3e-14
    tiny = 1e-300
    max_iterations = 300
    log_scale = -value + shape * math.log(value) - math.lgamma(shape)
    if value < shape + 1:
        term = 1.0 / shape
        total = term
        current_shape = shape
        for _ in range(max_iterations):
            current_shape += 1
            term *= value / current_shape
            total += term
            if abs(term) < abs(total) * epsilon:
                break
        lower = total * math.exp(log_scale)
        return min(1.0, max(0.0, 1.0 - lower))
    b = value + 1.0 - shape
    c = 1.0 / tiny
    d = 1.0 / max(b, tiny)
    fraction = d
    for iteration in range(1, max_iterations + 1):
        coefficient = -iteration * (iteration - shape)
        b += 2.0
        d = coefficient * d + b
        if abs(d) < tiny:
            d = tiny
        c = b + coefficient / c
        if abs(c) < tiny:
            c = tiny
        d = 1.0 / d
        delta = d * c
        fraction *= delta
        if abs(delta - 1.0) < epsilon:
            break
    return min(1.0, max(0.0, math.exp(log_scale) * fraction))


def _chi_square_sf(statistic: float, degrees_of_freedom: int) -> float:
    if statistic < 0 or degrees_of_freedom <= 0:
        raise ValueError("invalid chi-square arguments")
    return _regularized_gamma_q(degrees_of_freedom / 2.0, statistic / 2.0)


def _kruskal_wallis(samples: list[Any]) -> tuple[float, float]:
    np, pd = _libraries()
    arrays = [np.asarray(sample, dtype=float) for sample in samples if len(sample)]
    if len(arrays) < 2:
        raise ValueError("Kruskal-Wallis requires at least two groups")
    sizes = [len(sample) for sample in arrays]
    combined = np.concatenate(arrays)
    ranks = pd.Series(combined).rank(method="average").to_numpy(dtype=float)
    rank_sums: list[float] = []
    offset = 0
    for size in sizes:
        rank_sums.append(float(ranks[offset : offset + size].sum()))
        offset += size
    total = len(combined)
    statistic = 12.0 / (total * (total + 1.0)) * sum(total_rank**2 / size for total_rank, size in zip(rank_sums, sizes, strict=True)) - 3.0 * (total + 1.0)
    _, tie_counts = np.unique(combined, return_counts=True)
    correction = 1.0 - float(np.sum(tie_counts**3 - tie_counts)) / (total**3 - total) if total > 1 else 1.0
    if correction <= 0:
        raise ValueError("all observations are tied")
    statistic = max(0.0, statistic / correction)
    return statistic, _chi_square_sf(statistic, len(arrays) - 1)


def _chi_square_independence(table: Any) -> tuple[float, float]:
    np, _ = _libraries()
    observed = np.asarray(table, dtype=float)
    total = float(observed.sum())
    row_totals = observed.sum(axis=1)
    column_totals = observed.sum(axis=0)
    expected = np.outer(row_totals, column_totals) / total
    if np.any(expected <= 0):
        raise ValueError("chi-square expected counts must be positive")
    statistic = float(np.sum((observed - expected) ** 2 / expected))
    degrees_of_freedom = (observed.shape[0] - 1) * (observed.shape[1] - 1)
    return statistic, _chi_square_sf(statistic, degrees_of_freedom)


def _metric_statistics(
    frame: Any,
    pairs: list[dict[str, Any]],
    *,
    group_column: str,
    groups: list[str],
) -> list[dict[str, Any]]:
    np, pd = _libraries()
    group_series = frame[group_column].map(_clean_text)
    records: list[dict[str, Any]] = []
    for pair in pairs:
        metric = str(pair["metric"])
        values = pd.to_numeric(frame[pair["value_column"]], errors="coerce")
        if pair["flag_column"]:
            directions = frame[pair["flag_column"]].map(_flag_direction)
        else:
            directions = pd.Series("normal", index=frame.index)
        valid = values.notna()
        high = directions.eq("high") & valid
        low = directions.eq("low") & valid
        abnormal = high | low
        per_group: dict[str, Any] = {}
        numeric_samples: list[Any] = []
        contingency: list[list[int]] = []
        for group in groups:
            group_mask = group_series.eq(group) & valid
            sample = values[group_mask].dropna().astype(float)
            if len(sample) >= 2:
                numeric_samples.append(sample.to_numpy())
            group_n = int(group_mask.sum())
            group_high = int((group_mask & high).sum())
            group_low = int((group_mask & low).sum())
            group_abnormal = group_high + group_low
            contingency.append([group_abnormal, max(0, group_n - group_abnormal)])
            per_group[group] = {
                "n": group_n,
                "median": float(sample.median()) if group_n else None,
                "q1": float(sample.quantile(0.25)) if group_n else None,
                "q3": float(sample.quantile(0.75)) if group_n else None,
                "high_n": group_high,
                "low_n": group_low,
                "abnormal_n": group_abnormal,
                "abnormal_rate": float(group_abnormal / group_n) if group_n else None,
            }

        kruskal_p: float | None = None
        epsilon_squared: float | None = None
        if len(numeric_samples) >= 2:
            try:
                statistic, kruskal_p_raw = _kruskal_wallis(numeric_samples)
                kruskal_p = float(kruskal_p_raw)
                total_n = sum(len(sample) for sample in numeric_samples)
                epsilon_squared = float(max(0.0, (statistic - len(numeric_samples) + 1) / (total_n - len(numeric_samples))))
            except ValueError:
                pass

        chi2_p: float | None = None
        cramers_v: float | None = None
        table = np.asarray(contingency, dtype=float)
        if table.shape[0] >= 2 and table.sum() > 0 and np.all(table.sum(axis=1) > 0) and np.all(table.sum(axis=0) > 0):
            try:
                chi2, chi2_p_raw = _chi_square_independence(table)
                chi2_p = float(chi2_p_raw)
                cramers_v = float(math.sqrt(chi2 / table.sum()))
            except ValueError:
                pass

        numeric = values[valid].astype(float)
        unknown_flags = sorted({_clean_text(item) for item in frame[pair["flag_column"]].dropna().unique() if _flag_direction(item) == "unknown"}) if pair["flag_column"] else []
        record = {
            "metric": metric,
            "value_column": pair["value_column"],
            "flag_column": pair["flag_column"],
            "n": int(valid.sum()),
            "missing_n": int((~valid).sum()),
            "missing_rate": float((~valid).mean()),
            "median": float(numeric.median()) if len(numeric) else None,
            "q1": float(numeric.quantile(0.25)) if len(numeric) else None,
            "q3": float(numeric.quantile(0.75)) if len(numeric) else None,
            "high_n": int(high.sum()),
            "low_n": int(low.sum()),
            "abnormal_n": int(abnormal.sum()),
            "abnormal_rate": float(abnormal.sum() / valid.sum()) if valid.sum() else None,
            "unknown_flags": unknown_flags,
            "per_group": per_group,
            "tests": {
                "kruskal_wallis_p": kruskal_p,
                "kruskal_wallis_q": None,
                "epsilon_squared": epsilon_squared,
                "abnormality_chi_square_p": chi2_p,
                "abnormality_chi_square_q": None,
                "cramers_v": cramers_v,
            },
        }
        records.append(record)

    kruskal_q = _benjamini_hochberg([item["tests"]["kruskal_wallis_p"] for item in records])
    chi2_q = _benjamini_hochberg([item["tests"]["abnormality_chi_square_p"] for item in records])
    for record, numeric_q, abnormal_q in zip(records, kruskal_q, chi2_q, strict=True):
        record["tests"]["kruskal_wallis_q"] = numeric_q
        record["tests"]["abnormality_chi_square_q"] = abnormal_q
    return records


def _clinical_profile(
    source: Path,
    frame: Any,
    *,
    sheet_name: str | None,
    group_column: str,
    sex_column: str,
    age_column: str,
    diagnosis_column: str,
    metadata_columns: int,
    group_order: list[str] | None,
) -> tuple[dict[str, Any], Any]:
    _, pd = _libraries()
    required = [group_column, sex_column, age_column, diagnosis_column]
    missing = [column for column in required if column not in frame.columns]
    if missing:
        raise DataStudyError(f"required clinical column(s) missing: {missing}")
    pairs = _infer_metric_pairs(frame, metadata_columns)
    groups = _ordered_groups(frame[group_column], group_order)
    if len(groups) < 2:
        raise DataStudyError("clinical comparison requires at least two non-empty groups")
    metrics = _metric_statistics(frame, pairs, group_column=group_column, groups=groups)
    age_years = frame[age_column].map(_parse_age_years)
    group_series = frame[group_column].map(_clean_text)
    sex_series = frame[sex_column].map(_clean_text)
    cohorts: list[dict[str, Any]] = []
    for group in groups:
        mask = group_series.eq(group)
        ages = age_years[mask].dropna()
        cohorts.append(
            {
                "group": group,
                "display_label": GROUP_LABELS.get(group, group),
                "n": int(mask.sum()),
                "female_n": int((mask & sex_series.eq("女")).sum()),
                "male_n": int((mask & sex_series.eq("男")).sum()),
                "age_n": int(ages.notna().sum()),
                "age_median_years": float(ages.median()) if len(ages) else None,
                "age_q1_years": float(ages.quantile(0.25)) if len(ages) else None,
                "age_q3_years": float(ages.quantile(0.75)) if len(ages) else None,
            }
        )
    diagnoses = frame[diagnosis_column].map(_clean_text).replace("", pd.NA)
    top_diagnoses = [
        {"diagnosis": str(label), "n": int(count)}
        for label, count in diagnoses.value_counts().head(15).items()
    ]
    profile = {
        "schema_version": "0.1",
        "domain": "clinical-cbc",
        "source": {
            "path": str(source.resolve()),
            "sha256": sha256_file(source),
            "bytes": source.stat().st_size,
            "sheet": sheet_name,
        },
        "statistical_unit": "one workbook row; repeated-person status is unknown",
        "rows": int(len(frame)),
        "columns": int(len(frame.columns)),
        "group_column": group_column,
        "groups": groups,
        "cohorts": cohorts,
        "age": {
            "source_column": age_column,
            "conversion": "岁→years; 月→years/12; 天→years/365.25; unitless numeric values treated as years",
            "parsed_n": int(age_years.notna().sum()),
            "missing_n": int(age_years.isna().sum()),
            "median_years": float(age_years.median()),
            "q1_years": float(age_years.quantile(0.25)),
            "q3_years": float(age_years.quantile(0.75)),
            "min_years": float(age_years.min()),
            "max_years": float(age_years.max()),
        },
        "diagnoses": {
            "source_column": diagnosis_column,
            "non_missing_n": int(diagnoses.notna().sum()),
            "missing_n": int(diagnoses.isna().sum()),
            "unique_n": int(diagnoses.nunique()),
            "top": top_diagnoses,
        },
        "metrics": metrics,
        "interpretation_limits": [
            "The workbook does not supply measurement units or reference intervals.",
            "Abnormality uses only source-supplied arrow flags; no clinical threshold is inferred.",
            "The dataset is observational and heterogeneous; results are descriptive, not causal.",
            "Repeated measurements or repeated patients cannot be identified from the supplied fields.",
        ],
    }
    return profile, age_years


def _summary_frame(profile: dict[str, Any]) -> Any:
    _, pd = _libraries()
    rows: list[dict[str, Any]] = []
    for metric in profile["metrics"]:
        row: dict[str, Any] = {
            "metric": metric["metric"],
            "n": metric["n"],
            "missing_n": metric["missing_n"],
            "missing_rate": metric["missing_rate"],
            "median": metric["median"],
            "q1": metric["q1"],
            "q3": metric["q3"],
            "high_n": metric["high_n"],
            "low_n": metric["low_n"],
            "abnormal_rate": metric["abnormal_rate"],
            "kruskal_wallis_p": metric["tests"]["kruskal_wallis_p"],
            "kruskal_wallis_q": metric["tests"]["kruskal_wallis_q"],
            "epsilon_squared": metric["tests"]["epsilon_squared"],
            "abnormality_chi_square_p": metric["tests"]["abnormality_chi_square_p"],
            "abnormality_chi_square_q": metric["tests"]["abnormality_chi_square_q"],
            "cramers_v": metric["tests"]["cramers_v"],
        }
        for group in profile["groups"]:
            group_stats = metric["per_group"][group]
            row[f"{group}_n"] = group_stats["n"]
            row[f"{group}_median"] = group_stats["median"]
            row[f"{group}_abnormal_rate"] = group_stats["abnormal_rate"]
        rows.append(row)
    return pd.DataFrame(rows)


def _metric_map(profile: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {str(item["metric"]): item for item in profile["metrics"]}


def _format_q(value: float | None) -> str:
    if value is None or not math.isfinite(value):
        return "q=n/a"
    if value < 0.001:
        return "q<0.001"
    return f"q={value:.3f}"


def _render_clinical_figure(
    frame: Any,
    profile: dict[str, Any],
    output_dir: Path,
    *,
    hero_metric: str,
    title: str,
    dpi: int,
) -> tuple[dict[str, str], dict[str, Any]]:
    import matplotlib as mpl

    mpl.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.colors import Normalize
    from matplotlib.patches import Rectangle
    from matplotlib.transforms import Bbox

    _, pd = _libraries()
    metrics_by_name = _metric_map(profile)
    if hero_metric not in metrics_by_name:
        raise DataStudyError(f"hero metric not found: {hero_metric}")
    groups = list(profile["groups"])
    ordered_metrics = [name for name in METRIC_ORDER if name in metrics_by_name]
    ordered_metrics.extend(name for name in metrics_by_name if name not in ordered_metrics)
    group_series = frame[profile["group_column"]].map(_clean_text)

    rc = dict(PUBLICATION_RC)
    rc.update(
        {
            "axes.titlesize": 9.2,
            "axes.titleweight": 600,
            "font.size": 7.2,
            "savefig.bbox": None,
            "savefig.pad_inches": 0.0,
        }
    )
    with mpl.rc_context(rc):
        fig = plt.figure(figsize=(7.6, 8.2), constrained_layout=False)
        grid = fig.add_gridspec(
            3,
            2,
            left=0.10,
            right=0.97,
            top=0.855,
            bottom=0.08,
            width_ratios=(1.62, 1.0),
            height_ratios=(1.0, 1.02, 1.02),
            wspace=0.58,
            hspace=0.62,
        )
        ax_b = fig.add_subplot(grid[:, 0])
        ax_a = fig.add_subplot(grid[0, 1])
        ax_c = fig.add_subplot(grid[1, 1])
        ax_d = fig.add_subplot(grid[2, 1])

        matrix = [
            [float(metrics_by_name[metric]["per_group"][group]["abnormal_rate"] or 0.0) for group in groups]
            for metric in ordered_metrics
        ]
        vmax = max(0.35, min(0.75, max(max(row) for row in matrix)))
        norm = Normalize(vmin=0.0, vmax=vmax)
        cmap = mpl.colormaps["Blues"]
        for row_index, row in enumerate(matrix):
            for column_index, value in enumerate(row):
                color = cmap(norm(value))
                ax_b.add_patch(Rectangle((column_index, row_index), 1, 1, facecolor=color, edgecolor="white", linewidth=0.7))
                text_color = "white" if norm(value) > 0.58 else "#172027"
                ax_b.text(column_index + 0.5, row_index + 0.52, f"{value:.1%}", ha="center", va="center", fontsize=5.8, color=text_color)
        ax_b.set_xlim(0, len(groups))
        ax_b.set_ylim(len(ordered_metrics), 0)
        ax_b.set_xticks(
            [index + 0.5 for index in range(len(groups))],
            [GROUP_LABELS.get(group, group) for group in groups],
        )
        ax_b.set_yticks([index + 0.5 for index in range(len(ordered_metrics))], ordered_metrics)
        ax_b.xaxis.tick_top()
        ax_b.tick_params(axis="x", length=0, labeltop=True, labelbottom=False, pad=4)
        ax_b.tick_params(axis="y", length=0, pad=6)
        for spine in ax_b.spines.values():
            spine.set_visible(False)
        ax_b.text(
            0.0,
            1.045,
            "B  Source-flagged abnormality across care settings",
            transform=ax_b.transAxes,
            ha="left",
            va="bottom",
            fontsize=9.2,
            fontweight=600,
            clip_on=False,
        )
        scale_ax = ax_b.inset_axes([1.025, 0.15, 0.025, 0.68])
        steps = 24
        for step in range(steps):
            scale_ax.add_patch(
                Rectangle(
                    (0, step / steps),
                    1,
                    1 / steps,
                    facecolor=cmap(step / (steps - 1)),
                    edgecolor="none",
                )
            )
        scale_ax.set_xlim(0, 1)
        scale_ax.set_ylim(0, 1)
        scale_ax.set_xticks([])
        scale_ax.set_yticks([0, 0.5, 1], ["0%", f"{vmax / 2:.0%}", f"{vmax:.0%}"])
        scale_ax.yaxis.tick_right()
        scale_ax.set_ylabel("Flagged abnormality", rotation=270, labelpad=15, fontsize=7)
        scale_ax.yaxis.set_label_position("right")
        scale_ax.tick_params(axis="y", labelsize=6.5, length=2, pad=2)
        for spine in scale_ax.spines.values():
            spine.set_visible(False)

        cohort_order = list(reversed(groups))
        cohorts = {item["group"]: item for item in profile["cohorts"]}
        counts = [cohorts[group]["n"] for group in cohort_order]
        bars = ax_a.barh(
            range(len(cohort_order)),
            counts,
            color=[GROUP_COLORS.get(group, "#6C7A89") for group in cohort_order],
            height=0.62,
        )
        labels = []
        for group in cohort_order:
            cohort = cohorts[group]
            age = cohort["age_median_years"]
            q1 = cohort["age_q1_years"]
            q3 = cohort["age_q3_years"]
            labels.append(f"{GROUP_LABELS.get(group, group)}\nage {age:.0f} [{q1:.0f}–{q3:.0f}]")
        ax_a.set_yticks(range(len(cohort_order)), labels)
        ax_a.set_xlabel("Records")
        ax_a.set_title("A  Cohort composition", loc="left")
        ax_a.grid(axis="x", zorder=0)
        ax_a.grid(axis="y", visible=False)
        ax_a.spines[["top", "right", "left"]].set_visible(False)
        ax_a.tick_params(axis="y", length=0)
        for bar, count in zip(bars, counts, strict=True):
            ax_a.text(bar.get_width(), bar.get_y() + bar.get_height() / 2, f"  n={count:,}", va="center", fontsize=6.8)
        ax_a.set_xlim(0, max(counts) * 1.28)

        hero = metrics_by_name[hero_metric]
        values = pd.to_numeric(frame[hero["value_column"]], errors="coerce")
        positive_values = values[values > 0]
        use_log = bool(len(positive_values) and positive_values.quantile(0.99) / positive_values.quantile(0.01) >= 8)
        for group in groups:
            sample = values[group_series.eq(group) & values.notna()]
            if use_log:
                sample = sample[sample > 0]
            sample = sample.sort_values().to_numpy(dtype=float)
            if not len(sample):
                continue
            ecdf = (pd.Series(range(1, len(sample) + 1), dtype=float) / len(sample)).to_numpy()
            ax_c.step(sample, ecdf, where="post", label=GROUP_LABELS.get(group, group), color=GROUP_COLORS.get(group, "#6C7A89"), linewidth=1.35)
        if use_log:
            ax_c.set_xscale("log")
        ax_c.set_ylim(0, 1)
        ax_c.set_ylabel("Cumulative proportion")
        ax_c.set_xlabel(f"{hero_metric} (source units{' · log scale' if use_log else ''})")
        ax_c.set_title(f"C  Full {hero_metric} distribution", loc="left")
        ax_c.grid(True)
        ax_c.spines[["top", "right"]].set_visible(False)
        ax_c.legend(ncol=2, fontsize=6.2, loc="lower right", handlelength=1.6, columnspacing=0.8)

        association = sorted(
            (item for item in profile["metrics"] if item["tests"]["cramers_v"] is not None),
            key=lambda item: float(item["tests"]["cramers_v"]),
            reverse=True,
        )[:8]
        association.reverse()
        effect_values = [float(item["tests"]["cramers_v"]) for item in association]
        effect_labels = [str(item["metric"]) for item in association]
        effect_colors = ["#2A6F97" if (item["tests"]["abnormality_chi_square_q"] or 1) < 0.05 else "#AAB7C4" for item in association]
        effect_bars = ax_d.barh(range(len(association)), effect_values, color=effect_colors, height=0.62)
        ax_d.set_yticks(range(len(association)), effect_labels)
        ax_d.set_xlabel("Cramér's V")
        ax_d.set_title("D  Care-setting association", loc="left")
        ax_d.grid(axis="x")
        ax_d.grid(axis="y", visible=False)
        ax_d.spines[["top", "right", "left"]].set_visible(False)
        ax_d.tick_params(axis="y", length=0)
        maximum = max(effect_values) if effect_values else 0.1
        ax_d.set_xlim(0, max(0.14, maximum * 1.55))
        for bar, item in zip(effect_bars, association, strict=True):
            ax_d.text(bar.get_width(), bar.get_y() + bar.get_height() / 2, f"  {_format_q(item['tests']['abnormality_chi_square_q'])}", va="center", fontsize=6.2)

        fig.suptitle(title, x=0.10, y=0.97, ha="left", fontsize=13, fontweight=600, color="#172027")
        fig.text(
            0.10,
            0.94,
            f"{profile['rows']:,} records · {len(profile['metrics'])} CBC measures · abnormality defined only by source arrows",
            ha="left",
            va="top",
            fontsize=8.2,
            color="#42515b",
        )
        fig.text(
            0.10,
            0.007,
            "Exploratory aggregate analysis. Source workbook supplies no units or reference intervals. q values use Benjamini–Hochberg FDR; no causal or diagnostic claim is made.",
            ha="left",
            va="bottom",
            fontsize=6.3,
            color="#42515b",
        )

        fig.canvas.draw()
        renderer = fig.canvas.get_renderer()
        figure_width_points = fig.get_figwidth() * 72.0
        figure_height_points = fig.get_figheight() * 72.0
        display_to_points = 72.0 / fig.dpi

        def panel_bbox(axes: list[Any], padding: float = 5.0) -> list[float]:
            display_box = Bbox.union([axis.get_tightbbox(renderer) for axis in axes])
            left = max(0.0, display_box.x0 * display_to_points - padding)
            right = min(figure_width_points, display_box.x1 * display_to_points + padding)
            top = max(0.0, figure_height_points - display_box.y1 * display_to_points - padding)
            bottom = min(figure_height_points, figure_height_points - display_box.y0 * display_to_points + padding)
            return [round(left, 3), round(top, 3), round(right - left, 3), round(bottom - top, 3)]

        panel_manifest = {
            "panels": [
                {
                    "panel_id": "B",
                    "label": "Source-flagged abnormality heatmap across care settings",
                    "bbox": panel_bbox([ax_b, scale_ax], padding=3.0),
                },
                {
                    "panel_id": "A",
                    "label": "Cohort composition and age summary",
                    "bbox": panel_bbox([ax_a]),
                },
                {
                    "panel_id": "C",
                    "label": f"Full {hero_metric} empirical cumulative distributions",
                    "bbox": panel_bbox([ax_c]),
                },
                {
                    "panel_id": "D",
                    "label": "Care-setting association effect sizes",
                    "bbox": panel_bbox([ax_d]),
                },
            ]
        }

        svg = output_dir / "figure.svg"
        pdf = output_dir / "figure.pdf"
        preview = output_dir / "preview.png"
        fig.savefig(svg, format="svg", bbox_inches=None, pad_inches=0)
        fig.savefig(pdf, format="pdf", bbox_inches=None, pad_inches=0)
        fig.savefig(preview, format="png", dpi=dpi, bbox_inches=None, pad_inches=0)
        plt.close(fig)

    svg_audit = audit_svg(svg, allow_raster=False, require_text=True)
    if svg_audit.status != "PASS":
        raise DataStudyError(f"data figure SVG failed audit: {svg_audit.failures}")
    return (
        {"svg": str(svg.resolve()), "pdf": str(pdf.resolve()), "png": str(preview.resolve())},
        {"svg_audit": svg_audit.to_dict(), "panel_manifest": panel_manifest},
    )


def _write_markdown_artifacts(
    output_dir: Path,
    profile: dict[str, Any],
    source: Path,
    *,
    hero_metric: str,
    title: str,
) -> None:
    source_map = f"""# Data figure source map

- Source workbook: `{source.resolve()}`
- Source SHA-256: `{profile['source']['sha256']}`
- Sheet: `{profile['source']['sheet']}`
- Statistical unit: {profile['statistical_unit']}
- Plotting implementation: `figmirror.data_analysis.render_data_study`
- Derived analysis: `analysis_summary.csv`
- Main figure: `figure.svg`, `figure.pdf`, `preview.png`
- Panel exports: `panels/` (created by FigMirror finalization)

The source workbook is read-only. The figure and CSV contain aggregate results only.
"""
    (output_dir / "source_map.md").write_text(source_map, encoding="utf-8")
    design_brief = f"""# Figure design brief

## Figure identity

- Working title: {title}
- Primary reader question: How do source-flagged CBC abnormalities vary across care settings in this workbook?
- Intended conclusion: Care settings show distinct descriptive abnormality profiles, while the workbook alone does not support causal or diagnostic inference.
- Hero panel: B, the complete metric-by-setting abnormality atlas.

## Evidence chain

1. Panel A establishes cohort size and age composition.
2. Panel B shows all source-flagged abnormality rates without selecting only favorable metrics.
3. Panel C retains the full empirical distribution of the predeclared hero measure, {hero_metric}.
4. Panel D quantifies the care-setting association with Cramér's V and FDR-corrected tests.

## Scientific limits

- Source units and reference intervals are absent and are not inferred.
- Any arrow containing ↑ is normalized to high and any arrow containing ↓ to low.
- One row is the statistical unit; repeated-person status is unknown.
- Results are exploratory and aggregate-only.
"""
    (output_dir / "design_brief.md").write_text(design_brief, encoding="utf-8")


def render_data_study(
    source_path: str | Path,
    output_dir: str | Path,
    *,
    domain: str = "clinical-cbc",
    sheet_name: str | None = None,
    group_column: str = "患者类别",
    sex_column: str = "性别",
    age_column: str = "年龄",
    diagnosis_column: str = "临床诊断",
    metadata_columns: int = 5,
    group_order: list[str] | None = None,
    hero_metric: str = "WBC",
    title: str = "Care-setting patterns in routine complete blood counts",
    dpi: int = 300,
) -> dict[str, Any]:
    """Profile a tabular study and render a traceable publication figure.

    ``clinical-cbc`` is the first domain adapter.  The public dispatch point is
    intentionally generic so later domains can implement the same profile,
    design-brief, aggregate-table, figure, provenance, and QA contract.
    """

    if domain not in DATA_STUDY_DOMAINS:
        raise DataStudyError(f"unsupported data-study domain: {domain}")
    if metadata_columns < 1:
        raise DataStudyError("metadata_columns must be positive")
    if dpi <= 0:
        raise DataStudyError("dpi must be positive")
    source = Path(source_path)
    if not source.is_file():
        raise FileNotFoundError(source)
    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    raw_frame, resolved_sheet = _load_table(source, sheet_name)
    frame = _normalize_frame(raw_frame)
    profile, _ = _clinical_profile(
        source,
        frame,
        sheet_name=resolved_sheet,
        group_column=group_column,
        sex_column=sex_column,
        age_column=age_column,
        diagnosis_column=diagnosis_column,
        metadata_columns=metadata_columns,
        group_order=group_order,
    )
    summary = _summary_frame(profile)
    summary_path = output / "analysis_summary.csv"
    summary.to_csv(summary_path, index=False, encoding="utf-8-sig")
    profile_path = output / "data_profile.json"
    profile_path.write_text(json.dumps(profile, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    evidence_result = build_data_evidence_bundle(profile_path, output / "schematic_evidence.json")
    exports, render_evidence = _render_clinical_figure(
        frame,
        profile,
        output,
        hero_metric=hero_metric,
        title=title,
        dpi=dpi,
    )
    panel_manifest_path = output / "panel_manifest.json"
    panel_manifest_path.write_text(
        json.dumps(render_evidence["panel_manifest"], ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    data_binding = {
        "schema_version": "0.2",
        "source_files": [str(source.resolve())],
        "source_records": [profile["source"]],
        "sheet": resolved_sheet,
        "statistical_unit": profile["statistical_unit"],
        "grouping": {"column": group_column, "order": profile["groups"]},
        "metric_binding": [
            {
                "metric": item["metric"],
                "value_column": item["value_column"],
                "flag_column": item["flag_column"],
            }
            for item in profile["metrics"]
        ],
        "transformations": [
            "trim whitespace and line breaks without overwriting the workbook",
            "parse numeric measure columns with invalid tokens treated as missing",
            "normalize any ↑-containing flag to high and any ↓-containing flag to low",
            "convert ages expressed in years/months/days to years for cohort summaries",
            "compute group-wise medians, quartiles, and source-flagged abnormality rates",
            "apply Kruskal–Wallis and chi-square tests across care settings with Benjamini–Hochberg FDR",
        ],
        "units": "not supplied by source workbook; not inferred",
        "privacy": "figure and derived summary are aggregate-only",
    }
    (output / "data_binding.json").write_text(
        json.dumps(data_binding, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    figure_manifest = {
        "schema_version": "0.1",
        "figure_id": "medical-cbc-trial",
        "title": title,
        "scientific_question": "How do source-flagged CBC abnormalities vary across care settings?",
        "intended_conclusion": "Care settings have distinct descriptive abnormality profiles; causal and diagnostic inference is out of scope.",
        "hero_panel": "B",
        "panels": [
            {"id": "A", "question": "Who is represented?", "source": "data_profile.json", "transform": "counts and age median [IQR]"},
            {"id": "B", "question": "Which abnormalities vary by setting?", "source": "analysis_summary.csv", "transform": "source-flagged abnormal rate"},
            {"id": "C", "question": f"What is the full {hero_metric} distribution?", "source": str(source.resolve()), "transform": "empirical CDF; log x-axis only when spread warrants it"},
            {"id": "D", "question": "How strong is the setting association?", "source": "analysis_summary.csv", "transform": "chi-square Cramér's V and BH-FDR q"},
        ],
        "outputs": exports,
        "status": "generated-awaiting-visual-review",
    }
    (output / "figure_manifest.json").write_text(
        json.dumps(figure_manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    _write_markdown_artifacts(output, profile, source, hero_metric=hero_metric, title=title)
    qa = {
        "schema_version": "0.1",
        "status": "PASS_WITH_VISUAL_REVIEW_PENDING",
        "programmatic": {
            "source_hash_recorded": True,
            "source_rows": profile["rows"],
            "metric_pairs": len(profile["metrics"]),
            "svg_audit": render_evidence["svg_audit"],
            "outputs_nonempty": {name: Path(path).stat().st_size > 0 for name, path in exports.items()},
        },
        "scientific": {
            "source_thresholds_not_invented": True,
            "units_not_invented": True,
            "statistical_unit_disclosed": True,
            "multiple_comparisons_corrected": True,
            "aggregate_only_output": True,
        },
        "visual": {"status": "PENDING", "reviewer": None, "notes": []},
    }
    qa_path = output / "FIGURE_QA.json"
    qa_path.write_text(json.dumps(qa, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    result = {
        "schema_version": "0.1",
        "status": "PASS_WITH_VISUAL_REVIEW_PENDING",
        "domain": domain,
        "profile": str(profile_path.resolve()),
        "summary": str(summary_path.resolve()),
        "data_binding": str((output / "data_binding.json").resolve()),
        "figure_manifest": str((output / "figure_manifest.json").resolve()),
        "panel_manifest": str(panel_manifest_path.resolve()),
        "qa": str(qa_path.resolve()),
        "schematic_evidence": evidence_result["bundle"],
        "exports": exports,
    }
    (output / "data_study_report.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    return result
